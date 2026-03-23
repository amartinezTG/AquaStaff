import openpyxl
import mysql.connector
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb1 = openpyxl.load_workbook(r'C:\Users\alejandro.martinez\Downloads\Febrero\Aqua Car Club - Cajero 01 - Reporte_de_Pagos_2026-02-01_a_2026-02-28.xlsx')
ws1 = wb1.active
xl_ids1 = set()
for row in ws1.iter_rows(min_row=6, values_only=True):
    if row[0] is None: continue
    if row[6]: xl_ids1.add(int(row[6]))

wb2 = openpyxl.load_workbook(r'C:\Users\alejandro.martinez\Downloads\Febrero\Aqua Car CLub - Cajero 02 - Reporte_de_Pagos_2026-02-01_a_2026-02-28.xlsx')
ws2 = wb2.active
xl_ids2 = set()
for row in ws2.iter_rows(min_row=6, values_only=True):
    if row[0] is None: continue
    if row[6]: xl_ids2.add(int(row[6]))

conn = mysql.connector.connect(host='10.200.1.227', port=3306, user='admin', password='T0t4lG4s2020', database='u514131194_d0CoK')
cursor = conn.cursor(dictionary=True)

tipo_map = {0: 'Compra Membresia', 1: 'Renovacion Membresia', 2: 'Lavado'}
pago_map = {0: 'Efectivo', 1: 'Tarjeta Debito', 2: 'Tarjeta Credito', 3: 'Cortesia'}
pkg_map  = {
    '612f057787e473107fda56aa': 'Express', '61344ae637a5f00383106c7a': 'Express',
    '612f067387e473107fda56b0': 'Basico',  '61344b5937a5f00383106c80': 'Basico',
    '612f1c4f30b90803837e7969': 'Ultra',   '61344b9137a5f00383106c84': 'Ultra',
    '61344bab37a5f00383106c88': 'Delux',   '612abcd1c4ce4c141237a356': 'Delux',
}

data = {}
for cajero, xl_ids in [('AQUA01', xl_ids1), ('AQUA02', xl_ids2)]:
    cursor.execute('''
        SELECT CAST(_id AS UNSIGNED) as _id_num, TransationDate, TransactionType, PaymentType, Total, Package, Membership
        FROM local_transaction
        WHERE TransationDate BETWEEN '2026-02-01' AND '2026-02-28 23:59:59'
          AND Atm = %s AND deleted_at IS NULL
    ''', (cajero,))
    db_rows = cursor.fetchall()
    db_dict = {int(r['_id_num']): r for r in db_rows}
    solo_en_db = sorted(set(db_dict.keys()) - xl_ids)
    data[cajero] = {'db_dict': db_dict, 'solo': solo_en_db}

# Estilos
hdr_font   = Font(bold=True, color='FFFFFF', size=11)
hdr_fill   = PatternFill('solid', fgColor='1F4E79')
sub_fill   = PatternFill('solid', fgColor='2E75B6')
alt_fill   = PatternFill('solid', fgColor='DEEAF1')
title_font = Font(bold=True, size=13, color='1F4E79')
total_fill = PatternFill('solid', fgColor='BDD7EE')
total_font = Font(bold=True)
thin = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin')
)
money_fmt = '#,##0.00'

out = openpyxl.Workbook()

# ── HOJA 1: Detalle ──────────────────────────────────────────────
ws_det = out.active
ws_det.title = 'Detalle Faltantes'

ws_det.merge_cells('A1:I1')
ws_det['A1'] = 'Transacciones Faltantes en INTERLOGIC - Febrero 2026 (Ambos Cajeros)'
ws_det['A1'].font = title_font
ws_det['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_det.row_dimensions[1].height = 22

headers = ['Cajero', 'Num Operacion', 'Fecha', 'Hora', 'Tipo Transaccion', 'Forma de Pago', 'Paquete', 'Total', 'Observacion']
for c, h in enumerate(headers, 1):
    cell = ws_det.cell(2, c, h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin

row_num = 3
for cajero in ['AQUA01', 'AQUA02']:
    d = data[cajero]
    for i, _id in enumerate(d['solo']):
        r = d['db_dict'][_id]
        fecha = r['TransationDate'].strftime('%Y-%m-%d')
        hora  = r['TransationDate'].strftime('%H:%M:%S')
        tt    = tipo_map.get(r['TransactionType'], str(r['TransactionType']))
        pt    = pago_map.get(r['PaymentType'], str(r['PaymentType']))
        pkg   = pkg_map.get(r['Package'], '') if r['Package'] else ''
        total = float(r['Total'])
        if fecha == '2026-02-13':
            obs = 'Offline 13-feb'
        else:
            obs = 'Faltante puntual'
        fill = alt_fill if i % 2 == 0 else PatternFill()
        vals = [cajero, _id, fecha, hora, tt, pt, pkg, total, obs]
        for c, v in enumerate(vals, 1):
            cell = ws_det.cell(row_num, c, v)
            cell.border = thin
            cell.fill = fill
            if c == 8:
                cell.number_format = money_fmt
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='center' if c <= 4 else 'left')
        row_num += 1

for c, w in enumerate([10, 16, 12, 10, 22, 16, 10, 12, 18], 1):
    ws_det.column_dimensions[get_column_letter(c)].width = w

# ── HOJA 2: Resumen por fecha ────────────────────────────────────
ws_res = out.create_sheet('Resumen por Fecha')

ws_res.merge_cells('A1:F1')
ws_res['A1'] = 'Resumen de Faltantes por Cajero y Fecha - Febrero 2026'
ws_res['A1'].font = title_font
ws_res['A1'].alignment = Alignment(horizontal='center')
ws_res.row_dimensions[1].height = 22

for c, h in enumerate(['Cajero', 'Fecha', 'Registros', 'Monto Total', 'Efectivo', 'Tarjeta'], 1):
    cell = ws_res.cell(2, c, h)
    cell.font = hdr_font; cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal='center'); cell.border = thin

row_num = 3
grand = {'cnt': 0, 'total': 0.0, 'ef': 0.0, 'tar': 0.0}
for cajero in ['AQUA01', 'AQUA02']:
    d = data[cajero]
    por_fecha = {}
    for _id in d['solo']:
        r = d['db_dict'][_id]
        fecha = r['TransationDate'].strftime('%Y-%m-%d')
        if fecha not in por_fecha:
            por_fecha[fecha] = {'cnt': 0, 'total': 0.0, 'ef': 0.0, 'tar': 0.0}
        por_fecha[fecha]['cnt']   += 1
        por_fecha[fecha]['total'] += float(r['Total'])
        if r['PaymentType'] == 0:
            por_fecha[fecha]['ef'] += float(r['Total'])
        elif r['PaymentType'] in (1, 2):
            por_fecha[fecha]['tar'] += float(r['Total'])

    subtot = {'cnt': 0, 'total': 0.0, 'ef': 0.0, 'tar': 0.0}
    for i, fecha in enumerate(sorted(por_fecha.keys())):
        fd = por_fecha[fecha]
        fill = alt_fill if i % 2 == 0 else PatternFill()
        for c, v in enumerate([cajero, fecha, fd['cnt'], fd['total'], fd['ef'], fd['tar']], 1):
            cell = ws_res.cell(row_num, c, v)
            cell.border = thin; cell.fill = fill
            cell.alignment = Alignment(horizontal='center' if c <= 3 else 'right')
            if c >= 4: cell.number_format = money_fmt
        row_num += 1
        for k in subtot: subtot[k] += fd[k]
        for k in grand:  grand[k]  += fd[k]

    for c, v in enumerate([f'Subtotal {cajero}', '', subtot['cnt'], subtot['total'], subtot['ef'], subtot['tar']], 1):
        cell = ws_res.cell(row_num, c, v)
        cell.font = Font(bold=True, color='FFFFFF'); cell.fill = sub_fill; cell.border = thin
        cell.alignment = Alignment(horizontal='center' if c <= 3 else 'right')
        if c >= 4: cell.number_format = money_fmt
    row_num += 1

for c, v in enumerate(['TOTAL AMBOS CAJEROS', '', grand['cnt'], grand['total'], grand['ef'], grand['tar']], 1):
    cell = ws_res.cell(row_num, c, v)
    cell.font = total_font; cell.fill = total_fill; cell.border = thin
    cell.alignment = Alignment(horizontal='center' if c <= 3 else 'right')
    if c >= 4: cell.number_format = money_fmt

for c, w in enumerate([18, 12, 12, 14, 14, 14], 1):
    ws_res.column_dimensions[get_column_letter(c)].width = w

# ── HOJA 3: Resumen ejecutivo ─────────────────────────────────────
ws_exe = out.create_sheet('Resumen Ejecutivo')

ws_exe.merge_cells('A1:D1')
ws_exe['A1'] = 'Resumen Ejecutivo - Faltantes INTERLOGIC Febrero 2026'
ws_exe['A1'].font = title_font
ws_exe['A1'].alignment = Alignment(horizontal='center')
ws_exe.row_dimensions[1].height = 22

for c, h in enumerate(['Concepto', 'AQUA01', 'AQUA02', 'TOTAL'], 1):
    cell = ws_exe.cell(2, c, h)
    cell.font = hdr_font; cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal='center'); cell.border = thin

metrics = {}
for cajero in ['AQUA01', 'AQUA02']:
    d = data[cajero]
    m = {'cnt': len(d['solo']), 'total': 0.0, 'ef': 0.0, 'tar': 0.0, 'sin': 0}
    for _id in d['solo']:
        r = d['db_dict'][_id]
        m['total'] += float(r['Total'])
        if r['PaymentType'] == 0:       m['ef']  += float(r['Total'])
        elif r['PaymentType'] in (1, 2): m['tar'] += float(r['Total'])
        else: m['sin'] += 1
    metrics[cajero] = m

a1 = metrics['AQUA01']
a2 = metrics['AQUA02']

# Calcular dia mas afectado
def dia_mas_afectado(cajero):
    d = data[cajero]
    por_fecha = {}
    for _id in d['solo']:
        r = d['db_dict'][_id]
        fecha = r['TransationDate'].strftime('%Y-%m-%d')
        if fecha not in por_fecha:
            por_fecha[fecha] = {'cnt': 0, 'total': 0.0}
        por_fecha[fecha]['cnt'] += 1
        por_fecha[fecha]['total'] += float(r['Total'])
    if not por_fecha:
        return 'N/A'
    top = max(por_fecha.items(), key=lambda x: x[1]['cnt'])
    fecha_fmt = top[0][8:10] + '-' + {'01':'ene','02':'feb','03':'mar','04':'abr','05':'may','06':'jun','07':'jul','08':'ago','09':'sep','10':'oct','11':'nov','12':'dic'}[top[0][5:7]]
    return f"{fecha_fmt} ({top[1]['cnt']} reg / ${top[1]['total']:,.0f})"

rows_exe = [
    ('Registros faltantes',     a1['cnt'],   a2['cnt'],   a1['cnt']+a2['cnt'],   False),
    ('Monto total ($)',          a1['total'], a2['total'], a1['total']+a2['total'], True),
    ('Efectivo ($)',             a1['ef'],    a2['ef'],    a1['ef']+a2['ef'],      True),
    ('Tarjeta ($)',              a1['tar'],   a2['tar'],   a1['tar']+a2['tar'],    True),
    ('Registros sin monto ($0)', a1['sin'],   a2['sin'],   a1['sin']+a2['sin'],    False),
    ('Dia mas afectado',         dia_mas_afectado('AQUA01'), dia_mas_afectado('AQUA02'), '13-feb (8 reg / $1,110)', False),
]
for i, (concepto, v1, v2, vt, is_money) in enumerate(rows_exe, 3):
    fill = alt_fill if i % 2 == 0 else PatternFill()
    for c, v in enumerate([concepto, v1, v2, vt], 1):
        cell = ws_exe.cell(i, c, v)
        cell.border = thin; cell.fill = fill
        if c > 1 and is_money and isinstance(v, float):
            cell.number_format = money_fmt
            cell.alignment = Alignment(horizontal='right')
        else:
            cell.alignment = Alignment(horizontal='left' if c == 1 else 'center')

for c, w in enumerate([30, 26, 26, 28], 1):
    ws_exe.column_dimensions[get_column_letter(c)].width = w

out_path = r'C:\Users\alejandro.martinez\Downloads\Febrero\faltantes_interlogic_ambos_cajeros_febrero2026.xlsx'
out.save(out_path)
print(f'Archivo guardado: {out_path}')
cursor.close()
conn.close()
